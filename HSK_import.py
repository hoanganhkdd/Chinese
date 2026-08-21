"""
HSK_import.py — Nhập từ vựng HSK từ PDF vào Excel
===================================================
• Lần đầu: tạo file xlsx mới
• Lần sau: INSERT từ mới lên ĐẦU, giữ nguyên dữ liệu cũ
• Tự phát hiện trùng lặp, ghi chú vào cột GHI CHÚ
• Hỗ trợ: HSK1, HSK2, HSK3, HSK4 (bất kỳ PDF nào có format STT / Hán / Pinyin / Nghĩa)

Cách dùng:
  python HSK_import.py <file.pdf> [--output <file.xlsx>] [--level HSK1]

Ví dụ:
  python HSK_import.py HSK1_Tu_Vung_500.pdf
  python HSK_import.py HSK2_Tu_Vung.pdf --output HSK_TuVung_Master.xlsx --level HSK2
"""

import argparse
import re
import sys
from datetime import date
from pathlib import Path

# ── Third-party (install via: pip install pdfplumber openpyxl) ───────────────
try:
    import pdfplumber
except ImportError:
    sys.exit("Cài đặt: pip install pdfplumber")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Cài đặt: pip install openpyxl")

# ── Constants ────────────────────────────────────────────────────────────────
TODAY = date.today().strftime("%Y-%m-%d")

HEADERS = [
    ("STT",              4),
    ("CẤP ĐỘ",           8),
    ("🀄 HÁN TỰ",       12),
    ("🔊 PINYIN",        14),
    ("🇻🇳 TIẾNG VIỆT",  22),
    ("📝 VIẾT THỬ 1",   16),
    ("📝 VIẾT THỬ 2",   16),
    ("✅ ĐÁP ÁN",       12),
    ("🎵 AUDIO",         14),
    ("🌐 YOUGLISH",      14),
    ("DAY 1",  8), ("DAY 2",  8), ("DAY 3",  8), ("DAY 4",  8),
    ("DAY 5",  8), ("DAY 6",  8), ("DAY 7",  8),
    ("📅 NGÀY THÊM",     13),
    ("⚠️ GHI CHÚ",      28),
]
COL = {h[0]: i + 1 for i, h in enumerate(HEADERS)}
N_COLS = len(HEADERS)

# Colors
C = {
    "header_bg": "B71C1C", "header_fg": "FFFFFF",
    "hanzi_bg":  "FFCDD2", "pinyin_bg": "E8F5E9",
    "viet_bg":   "E3F2FD", "input_bg":  "FFFDE7",
    "ans_bg":    "F3E5F5", "link_bg":   "E8EAF6",
    "day_odd":   "FCE4EC", "day_even":  "F8BBD9",
    "date_bg":   "FFF9C4", "note_bg":   "FFE0B2",
    "alt1":      "FFEBEE", "alt2":      "FFFFFF",
    "dup_note":  "FF6F00",
}

LEVEL_BG = {
    "HSK1": "E53935", "HSK2": "8E24AA", "HSK3": "1E88E5",
    "HSK4": "00897B", "HSK5": "F4511E", "HSK6": "6D4C41",
}

thin = Side(style="thin", color="BDBDBD")
medium = Side(style="medium", color="9E9E9E")


# ── Helpers ──────────────────────────────────────────────────────────────────
def border(light=True):
    s = thin if light else medium
    return Border(left=s, right=s, top=s, bottom=s)


def style(cell, bold=False, bg=None, fg="000000", size=10,
          h="center", v="center", wrap=False):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def make_hyperlink(url, label):
    return f'=HYPERLINK("{url}","{label}")'


# ── PDF extraction ───────────────────────────────────────────────────────────
PATTERN = re.compile(
    r'^(\d+)\s+'
    r'([一-鿿㐀-䶿豈-﫿]+)\s+'
    r'([a-zA-ZÀ-ỹāáǎàēéěèīíǐìōóǒòūúǔùǖǘǚǜü\s,·]+?)\s+'
    r'(.+)$'
)


def extract_from_pdf(pdf_path: Path):
    words = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            if i == 0:
                continue
            text = page.extract_text() or ""
            for line in text.split("\n"):
                m = PATTERN.match(line.strip())
                if m:
                    words.append({
                        "stt":    int(m.group(1)),
                        "hanzi":  m.group(2).strip(),
                        "pinyin": m.group(3).strip(),
                        "viet":   m.group(4).strip(),
                    })
    return words


# ── Excel read/write ─────────────────────────────────────────────────────────
def load_existing(xlsx_path: Path):
    """Return (wb, ws, existing_hanzi_set)"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    existing = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        hanzi = row[COL["🀄 HÁN TỰ"] - 1]
        if hanzi:
            existing.add(str(hanzi).strip())
    return wb, ws, existing


def write_header_row(ws, row_num):
    for col_i, (hdr, _) in enumerate(HEADERS, start=1):
        c = ws.cell(row=row_num, column=col_i, value=hdr)
        style(c, bold=True, bg=C["header_bg"], fg=C["header_fg"], size=10)
        c.border = border(light=False)


def write_data_row(ws, row_num, word, level, is_dup=False, dup_note=""):
    hanzi  = word["hanzi"]
    pinyin = word["pinyin"]
    viet   = word["viet"]
    row_bg = C["alt1"] if row_num % 2 == 0 else C["alt2"]

    lvl_bg = LEVEL_BG.get(level.upper(), "616161")

    values = {
        "STT":              "",                 # renumbered after insert
        "CẤP ĐỘ":           level.upper(),
        "🀄 HÁN TỰ":        hanzi,
        "🔊 PINYIN":         pinyin,
        "🇻🇳 TIẾNG VIỆT":   viet,
        "📝 VIẾT THỬ 1":    "",
        "📝 VIẾT THỬ 2":    "",
        "✅ ĐÁP ÁN":        "",
        "🎵 AUDIO":          make_hyperlink(
            f"https://translate.google.com/translate_tts?ie=UTF-8&tl=zh-CN&client=tw-ob&q={hanzi}",
            "▶"
        ),
        "🌐 YOUGLISH":       make_hyperlink(
            f"https://youglish.com/pronounce/{hanzi}/chinese",
            "🔍"
        ),
        "📅 NGÀY THÊM":     TODAY,
        "⚠️ GHI CHÚ":       dup_note if is_dup else "",
    }

    for col_name, val in values.items():
        c = ws.cell(row=row_num, column=COL[col_name], value=val)
        if col_name == "STT":
            style(c, bg=C["header_bg"], fg=C["header_fg"], bold=True, size=9)
        elif col_name == "CẤP ĐỘ":
            style(c, bg=lvl_bg, fg="FFFFFF", bold=True, size=9)
        elif col_name == "🀄 HÁN TỰ":
            style(c, bg=C["hanzi_bg"], size=14, bold=True)
        elif col_name == "🔊 PINYIN":
            style(c, bg=C["pinyin_bg"])
        elif col_name == "🇻🇳 TIẾNG VIỆT":
            style(c, bg=C["viet_bg"], h="left", wrap=True)
        elif col_name in ("📝 VIẾT THỬ 1", "📝 VIẾT THỬ 2"):
            style(c, bg=C["input_bg"])
        elif col_name == "✅ ĐÁP ÁN":
            style(c, bg=C["ans_bg"])
        elif col_name in ("🎵 AUDIO", "🌐 YOUGLISH"):
            c.font = Font(name="Arial", color="1565C0", size=12, underline="single")
            c.fill = PatternFill("solid", start_color=C["link_bg"])
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_name == "📅 NGÀY THÊM":
            style(c, bg=C["date_bg"], size=9)
        elif col_name == "⚠️ GHI CHÚ":
            style(c, bg=C["note_bg"] if is_dup else row_bg,
                  fg=C["dup_note"] if is_dup else "000000", h="left", wrap=True)
        else:
            style(c, bg=row_bg)
        c.border = border()

    # DAY columns
    for day in range(1, 8):
        c = ws.cell(row=row_num, column=COL["DAY 1"] + day - 1, value="")
        style(c, bg=C["day_odd"] if day % 2 == 1 else C["day_even"])
        c.border = border()

    ws.row_dimensions[row_num].height = 20


def renumber_stt(ws, data_start=3):
    """Rewrite STT column 1,2,3,... from top"""
    stt = 1
    for row in ws.iter_rows(min_row=data_start):
        hanzi_cell = row[COL["🀄 HÁN TỰ"] - 1]
        if hanzi_cell.value:
            stt_cell = row[COL["STT"] - 1]
            stt_cell.value = stt
            stt += 1


def create_fresh(xlsx_path: Path, words, level):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "📚 Từ Vựng HSK"

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    tc = ws["A1"]
    tc.value = f"🀄 TỪ VỰNG {level.upper()} — {len(words)} TỪ | Hệ thống học 7 ngày"
    style(tc, bold=True, bg="B71C1C", fg="FFFFFF", size=14)
    ws.row_dimensions[1].height = 30

    # Header
    write_header_row(ws, 2)
    ws.row_dimensions[2].height = 22

    # Set column widths
    for col_i, (_, w) in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.freeze_panes = "A3"

    for i, word in enumerate(words):
        write_data_row(ws, i + 3, word, level)

    renumber_stt(ws)
    wb.save(xlsx_path)
    return len(words), 0


def append_to_existing(xlsx_path: Path, words, level, existing_hanzi):
    wb, ws, _ = load_existing(xlsx_path)

    new_words = []
    dup_words = []
    for w in words:
        if w["hanzi"] in existing_hanzi:
            dup_words.append(w)
        else:
            new_words.append(w)

    if not new_words and not dup_words:
        print("Không có từ nào để thêm.")
        return 0, 0

    # Update title row
    tc = ws["A1"]
    # Count existing data rows
    existing_count = sum(
        1 for row in ws.iter_rows(min_row=3, values_only=True)
        if row[COL["🀄 HÁN TỰ"] - 1]
    )
    total = existing_count + len(new_words)
    tc.value = f"🀄 TỪ VỰNG HSK TỔNG HỢP — {total} TỪ | Cập nhật {TODAY}"

    # Mark duplicates in existing rows
    for dup in dup_words:
        for row in ws.iter_rows(min_row=3):
            hanzi_cell = row[COL["🀄 HÁN TỰ"] - 1]
            if hanzi_cell.value == dup["hanzi"]:
                note_cell = row[COL["⚠️ GHI CHÚ"] - 1]
                existing_note = note_cell.value or ""
                new_note = f"⚠️ Trùng với {level.upper()} ({TODAY})"
                note_cell.value = (existing_note + " | " + new_note).lstrip(" | ")
                note_cell.font = Font(name="Arial", color=C["dup_note"], size=9, bold=True)
                note_cell.fill = PatternFill("solid", start_color=C["note_bg"])
                break

    # Insert new words at top (row 3), pushing existing down
    if new_words:
        ws.insert_rows(3, amount=len(new_words))
        for i, word in enumerate(new_words):
            write_data_row(ws, 3 + i, word, level)

    renumber_stt(ws)
    wb.save(xlsx_path)
    return len(new_words), len(dup_words)


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description="Nhập từ vựng HSK từ PDF vào Excel")
    p.add_argument("pdf", help="Đường dẫn file PDF")
    p.add_argument("--output", "-o", default=None,
                   help="File Excel đầu ra (mặc định: tên PDF + .xlsx)")
    p.add_argument("--level", "-l", default=None,
                   help="Cấp độ HSK (HSK1/HSK2/...); tự phát hiện nếu bỏ trống")
    args = p.parse_args()

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        sys.exit(f"Không tìm thấy file: {pdf_path}")

    # Auto-detect level from filename
    level = args.level
    if not level:
        m = re.search(r'HSK\s*(\d)', pdf_path.name, re.IGNORECASE)
        level = f"HSK{m.group(1)}" if m else "HSK"

    # Output path
    out_path = Path(args.output) if args.output else pdf_path.with_suffix(".xlsx")

    print(f"📄 PDF    : {pdf_path.name}")
    print(f"📊 Output : {out_path}")
    print(f"🏷️  Level  : {level.upper()}")
    print("─" * 45)

    # Extract
    print("⏳ Đang đọc PDF...", end=" ", flush=True)
    words = extract_from_pdf(pdf_path)
    print(f"{len(words)} từ tìm thấy")

    if not words:
        sys.exit("Không trích xuất được từ nào. Kiểm tra định dạng PDF.")

    # Write
    if out_path.exists():
        print(f"📂 File đã tồn tại — thêm từ mới lên đầu (top-on-top)...")
        _, existing_hanzi, _ = load_existing(out_path)[1], load_existing(out_path)[2], None
        wb2, ws2, existing_hanzi = load_existing(out_path)
        wb2.close() if hasattr(wb2, 'close') else None
        n_new, n_dup = append_to_existing(out_path, words, level, existing_hanzi)
        print(f"✅ Thêm mới : {n_new} từ")
        print(f"⚠️  Trùng lặp: {n_dup} từ (đã ghi chú)")
    else:
        print("🆕 Tạo file mới...")
        n_new, n_dup = create_fresh(out_path, words, level)
        print(f"✅ Tạo xong : {n_new} từ")

    print(f"\n📁 Lưu tại: {out_path.resolve()}")


if __name__ == "__main__":
    main()
