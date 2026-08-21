#!/usr/bin/env python3
"""
HA.py — YouTube Vocabulary Extractor for Chinese Learning
==========================================================
Inspired by NotebookLM — feed a YouTube video, get a study-ready
vocabulary xlsx mapped to your Vocab (TuVung_Trung-Anh-Viet) database.

WHAT IT DOES
------------
1. Fetch Chinese transcript from a YouTube video
2. Segment text into words using jieba
3. Map every word against your Vocab (HSK 1–4 + Business) xlsx
4. Extract TOP 100 words by frequency (configurable)
5. Output xlsx with 4 sheets — OR append to a master accumulation file:
      📹 Video       — summary & stats
      📚 Từ Vựng HA  — all words in Vocab study format (with source info)
      ✅ Đã Học      — words already in your Vocab
      🆕 Từ Mới      — new words not yet in Vocab

NEW IN v2.0
-----------
  • --top N          Only extract top N words by frequency (default: 100)
  • --master FILE    Accumulate all videos into one master xlsx:
                       - New words added to TOP (most recent first)
                       - Duplicate words NOT re-added; note updated instead
  • Source columns   Every row now records: video name, video link, date added
  • Duplicate notes  ⚠️ GHI CHÚ column shows which other videos contain the word

USAGE
-----
    python HA.py "https://www.youtube.com/watch?v=VIDEO_ID"
    python HA.py VIDEO_ID --vocab "./TuVung.xlsx" --top 50
    python HA.py VIDEO_ID --master "./HA_master.xlsx"   ← accumulate mode
    python HA.py VIDEO_ID --min-freq 2 --no-stopwords --output "./output"
    python HA.py VIDEO_ID --lang zh-TW

INSTALL DEPENDENCIES
--------------------
    pip install youtube-transcript-api jieba pypinyin openpyxl requests

Version: 2.0  |  Author: HA Module
"""

import sys
import re
import argparse
import unicodedata
import random
from pathlib import Path
from collections import Counter
from datetime import datetime

# ─── Dependency Check ──────────────────────────────────────────────────────────

MISSING = []

try:
    from youtube_transcript_api import YouTubeTranscriptApi
    from youtube_transcript_api._errors import TranscriptsDisabled, NoTranscriptFound
except ImportError:
    MISSING.append("youtube-transcript-api")

try:
    import jieba
except ImportError:
    MISSING.append("jieba")

try:
    from pypinyin import pinyin as to_pinyin, Style
except ImportError:
    MISSING.append("pypinyin")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    MISSING.append("openpyxl")

try:
    import requests
except ImportError:
    MISSING.append("requests")

if MISSING:
    print("❌  Missing dependencies. Run:")
    print(f"    pip install {' '.join(MISSING)}")
    sys.exit(1)

# ─── Constants ─────────────────────────────────────────────────────────────────

# YouTube transcript language preference order (Chinese variants)
CHINESE_LANG_CODES = [
    "zh-Hans", "zh-Hant", "zh", "zh-CN", "zh-TW",
    "zh-HK", "zh-SG", "zh-MO", "cmn", "yue",
]

# Vocab sheet names (must match the actual xlsx tab names exactly)
VOCAB_SHEETS = ["HSK 1", "HSK 2", "HSK 3", "HSK 4", "💼 Thương mại"]

# Column positions in Vocab xlsx (0-based, values_only row)
# STT | 🔊 NGHE TỪ | ENGLISH | TIẾNG VIỆT | ✍️ GÕ HÁN | ✅ KT HÁN |
# 🅰 GÕ PINYIN | ✅ KT PINYIN | 📖 ĐÁP ÁN HÁN | 🔤 PINYIN | 📝 VÍ DỤ |
# 🔊 NGHE VÍ DỤ | 🎬 YouGlish | 🔍 CHIẾT TỰ | ...
VC = {
    "stt": 0, "english": 2, "vietnamese": 3,
    "hanzi": 8, "pinyin": 9, "example": 10, "chiettu": 13,
}

# Common function words safe to remove (particles / interjections)
STOPWORDS = {
    "的", "了", "在", "和", "也", "就", "都", "着", "过",
    "吗", "呢", "啊", "吧", "嗯", "哦", "哈", "嘛", "呀",
    "哇", "哎", "嘿", "喂", "啦", "咯", "哟", "诶",
    "之", "以", "其", "而", "于", "与", "被", "把", "让",
}

# Background colours per HSK level
LEVEL_BG = {
    "HSK 1":          "E3F2FD",   # light blue
    "HSK 2":          "F3E5F5",   # light purple
    "HSK 3":          "FFF8E1",   # light yellow
    "HSK 4":          "FCE4EC",   # light pink
    "💼 Thương mại":  "E0F7FA",   # light cyan
    "🆕 Mới":         "FFF3E0",   # light orange
}

HEADER_DARK  = "1A3C5E"   # deep navy
HEADER_LIGHT = "FFFFFF"
VIDEO_DARK   = "263238"   # near-black
ACCENT_BLUE  = "1976D2"


# ─── Helpers ───────────────────────────────────────────────────────────────────

def is_chinese_char(c: str) -> bool:
    return '一' <= c <= '鿿' or '㐀' <= c <= '䶿'


def contains_chinese(word: str) -> bool:
    return any(is_chinese_char(c) for c in word)


def get_pinyin_str(word: str) -> str:
    """Return toned pinyin string, e.g. 'nǐ hǎo'."""
    try:
        result = to_pinyin(word, style=Style.TONE)
        return ' '.join(p[0] for p in result)
    except Exception:
        return ''


def audio_url(word: str) -> str:
    """Google Translate link for listening to a word."""
    return f"https://translate.google.com/?sl=zh-CN&tl=vi&text={word}&op=translate"


def youglish_url(word: str) -> str:
    return f"https://youglish.com/pronounce/{word}/chinese"


def safe_filename(name: str, max_len: int = 90) -> str:
    name = re.sub(r'[<>:"/\\|?*\n\r\t]', '_', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name[:max_len] if len(name) > max_len else name


# ─── Video Metadata ────────────────────────────────────────────────────────────

def extract_video_id(url: str) -> str:
    """Parse video ID from any standard YouTube URL or bare ID."""
    patterns = [
        r'(?:v=|/v/|youtu\.be/|/embed/)([A-Za-z0-9_-]{11})',
        r'^([A-Za-z0-9_-]{11})$',
    ]
    for p in patterns:
        m = re.search(p, url.strip())
        if m:
            return m.group(1)
    raise ValueError(f"Cannot extract YouTube video ID from: {url!r}")


def get_video_title(video_id: str) -> str:
    """Fetch video title via YouTube oEmbed (no API key needed)."""
    try:
        url = (
            "https://www.youtube.com/oembed"
            f"?url=https://www.youtube.com/watch?v={video_id}&format=json"
        )
        resp = requests.get(url, timeout=10)
        if resp.status_code == 200:
            return resp.json().get("title", f"Video_{video_id}")
    except Exception:
        pass
    return f"Video_{video_id}"


# ─── Transcript ────────────────────────────────────────────────────────────────

def fetch_transcript(video_id: str, force_lang: str = None) -> tuple:
    """
    Fetch Chinese transcript.  Returns (segments_list, lang_code_used).
    Tries: manual → auto-generated → translated from English.
    """
    lang_order = ([force_lang] + CHINESE_LANG_CODES) if force_lang else CHINESE_LANG_CODES

    try:
        transcript_list = YouTubeTranscriptApi.list_transcripts(video_id)
    except TranscriptsDisabled:
        raise RuntimeError("Transcripts are disabled for this video.")
    except Exception as e:
        raise RuntimeError(f"Cannot list transcripts: {e}")

    # 1. Manual transcripts (highest quality)
    for lang in lang_order:
        try:
            t = transcript_list.find_manually_created_transcript([lang])
            segs = t.fetch()
            print(f"   ✅ Manual transcript found: {lang}")
            return segs, lang
        except Exception:
            continue

    # 2. Auto-generated transcripts
    for lang in lang_order:
        try:
            t = transcript_list.find_generated_transcript([lang])
            segs = t.fetch()
            print(f"   ✅ Auto-generated transcript found: {lang}")
            return segs, lang
        except Exception:
            continue

    # 3. Translate English → Simplified Chinese
    try:
        t = transcript_list.find_transcript(['en'])
        t_zh = t.translate('zh-Hans')
        segs = t_zh.fetch()
        print("   ✅ Translated transcript (en → zh-Hans)")
        return segs, 'zh-Hans (translated)'
    except Exception:
        pass

    raise RuntimeError(
        "No Chinese transcript found. "
        "The video may not have captions, or they may be in another language. "
        "Try --lang en to translate from English."
    )


def segments_to_text(segments: list) -> str:
    """Flatten transcript segments into a single clean string."""
    parts = []
    for seg in segments:
        text = seg.get('text', '')
        text = re.sub(r'<[^>]+>', '', text)                  # strip HTML tags
        text = re.sub(r'[\[【\(（][^\]】\)）]*[\]】\)）]', '', text)  # remove [annotations]
        parts.append(text.strip())
    return ' '.join(parts)


# ─── Chinese Word Segmentation ─────────────────────────────────────────────────

def segment_chinese(text: str, min_chars: int = 1) -> Counter:
    """
    Segment Chinese text with jieba and return {word: frequency}.
    Only keeps tokens that contain at least one Chinese character
    and meet the min_chars length requirement.
    """
    freq: Counter = Counter()
    for word in jieba.cut(text, cut_all=False):
        word = word.strip()
        if not contains_chinese(word):
            continue
        if len(word) < min_chars:
            continue
        freq[word] += 1
    return freq


# ─── Vocab Database ────────────────────────────────────────────────────────────

def load_vocab(vocab_path: str) -> dict:
    """
    Load Vocab xlsx into a dict: {hanzi_string: entry_dict}.
    entry_dict keys: level, english, vietnamese, pinyin, example, chiettu
    """
    db: dict = {}
    if not vocab_path:
        return db

    path = Path(vocab_path)
    if not path.exists():
        print(f"   ⚠️  Vocab file not found: {vocab_path}")
        return db

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception as e:
        print(f"   ⚠️  Could not open Vocab file: {e}")
        return db

    for sheet in VOCAB_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[VC["stt"]]:
                continue
            try:
                int(row[VC["stt"]])
            except (TypeError, ValueError):
                continue
            hanzi = str(row[VC["hanzi"]] or "").strip()
            if not hanzi or not contains_chinese(hanzi):
                continue
            db[hanzi] = {
                "level":      sheet,
                "english":    str(row[VC["english"]]   or "").strip(),
                "vietnamese": str(row[VC["vietnamese"]] or "").strip(),
                "pinyin":     str(row[VC["pinyin"]]    or "").strip(),
                "example":    str(row[VC["example"]]   or "").strip(),
                "chiettu":    str(row[VC["chiettu"]]   or "").strip(),
            }
    wb.close()
    print(f"   ✅ Loaded {len(db):,} words from Vocab")
    return db


def auto_detect_vocab(script_dir: Path) -> str | None:
    """Search common locations for the Vocab xlsx."""
    search_dirs = [Path("."), script_dir]
    keywords = ("tuvung", "vocab", "chinese")
    for d in search_dirs:
        for f in d.glob("*.xlsx"):
            if any(k in f.name.lower() for k in keywords):
                return str(f)
    return None


# ─── Mapping ───────────────────────────────────────────────────────────────────

def map_to_vocab(
    word_freq: Counter,
    vocab_db: dict,
    remove_stopwords: bool = False,
) -> tuple[list, list]:
    """
    Partition words into (in_vocab, new_words).
    Each entry is a dict with: hanzi, freq, level, english, vietnamese,
                               pinyin, example, chiettu
    """
    level_order = {name: i for i, name in enumerate(VOCAB_SHEETS)}
    in_vocab, new_words = [], []

    for word, freq in word_freq.most_common():
        if remove_stopwords and word in STOPWORDS:
            continue
        if word in vocab_db:
            entry = dict(vocab_db[word])
            entry.update(hanzi=word, freq=freq)
            in_vocab.append(entry)
        else:
            new_words.append({
                "hanzi":      word,
                "freq":       freq,
                "level":      "🆕 Mới",
                "english":    "",
                "vietnamese": "",
                "pinyin":     get_pinyin_str(word),
                "example":    "",
                "chiettu":    "",
            })

    in_vocab.sort(key=lambda x: (level_order.get(x["level"], 99), -x["freq"]))
    new_words.sort(key=lambda x: -x["freq"])
    return in_vocab, new_words


# ─── Excel Helpers ─────────────────────────────────────────────────────────────

def _hdr(cell, bg=HEADER_DARK, fg=HEADER_LIGHT, bold=True, size=10):
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _dat(cell, bg=None, bold=False, wrap=True, align="left", size=10, color="000000"):
    if bg:
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.font      = Font(name="Calibri", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


def _link(ws, row, col, url, label, bg=None, fg=ACCENT_BLUE):
    c = ws.cell(row=row, column=col, value=label)
    c.hyperlink = url
    c.font      = Font(name="Calibri", color=fg, underline="single", size=10)
    if bg:
        c.fill = PatternFill(fill_type="solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ─── Master File (accumulation) ───────────────────────────────────────────────

def load_master_data(master_path: str) -> tuple[dict, list]:
    """
    Read an existing master xlsx and return:
      existing_hanzi : {hanzi: {"source_video": ..., "note": ...}}
      existing_words : list of word dicts (same structure as in_vocab / new_words)
    Returns ({}, []) if the file does not exist yet.
    """
    path = Path(master_path)
    if not path.exists():
        return {}, []

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception as e:
        print(f"   ⚠️  Cannot read master file: {e}")
        return {}, []

    # Find the first vocab sheet (📚 Từ Vựng HA or the first sheet)
    target_sheet = None
    for name in ["📚 Từ Vựng HA", "✅ Đã Học", "🆕 Từ Mới"]:
        if name in wb.sheetnames:
            target_sheet = wb[name]
            break
    if target_sheet is None and wb.sheetnames:
        target_sheet = wb[wb.sheetnames[0]]

    existing_hanzi = {}
    existing_words = []

    if target_sheet:
        mc = MASTER_COL
        for row in target_sheet.iter_rows(min_row=4, values_only=True):
            if not row or not row[mc["stt"]]:
                continue
            try:
                int(row[mc["stt"]])
            except (TypeError, ValueError):
                continue
            hanzi = str(row[mc["hanzi"]] or "").strip()
            if not hanzi or not contains_chinese(hanzi):
                continue
            wd = {
                "hanzi":        hanzi,
                "english":      str(row[mc["english"]]      or "").strip(),
                "vietnamese":   str(row[mc["vietnamese"]]   or "").strip(),
                "pinyin":       str(row[mc["pinyin"]]       or "").strip(),
                "example":      str(row[mc["example"]]      or "").strip(),
                "chiettu":      str(row[mc["chiettu"]]      or "").strip(),
                "freq":         int(row[mc["freq"]]  or 0) if row[mc["freq"]] else 0,
                "level":        str(row[mc["level"]]        or "").strip(),
                "source_video": str(row[mc["source_video"]] or "").strip(),
                "source_link":  str(row[mc["source_link"]]  or "").strip(),
                "added_date":   str(row[mc["added_date"]]   or "").strip(),
                "note":         str(row[mc["note"]]         or "").strip(),
            }
            existing_hanzi[hanzi] = {
                "source_video": wd["source_video"],
                "note":         wd["note"],
            }
            existing_words.append(wd)

    wb.close()
    print(f"   ✅ Master file: {len(existing_words)} existing words")
    return existing_hanzi, existing_words


def merge_with_master(
    new_words: list,
    existing_hanzi: dict,
    existing_words: list,
    video_title: str,
    video_url: str,
) -> tuple[list, int, int]:
    """
    Merge newly extracted words with the existing master data.

    - New words  → prepend with source info  (go to TOP)
    - Duplicates → NOT re-added; instead append a note to the existing row

    Returns:
        combined      : [new_words_with_source] + [existing_words_with_updated_notes]
        n_new         : count of genuinely new words added
        n_duplicates  : count of duplicates (notes updated)
    """
    today    = datetime.now().strftime("%Y-%m-%d")
    tag      = f"⚠️ Cũng trong: {video_title} ({today})"
    n_new    = 0
    n_dupes  = 0
    truly_new = []

    for wd in new_words:
        hanzi = wd["hanzi"]
        if hanzi in existing_hanzi:
            # Update note on the existing row
            for ex_wd in existing_words:
                if ex_wd["hanzi"] == hanzi:
                    current_note = ex_wd.get("note", "")
                    if tag not in current_note:
                        ex_wd["note"] = (
                            (current_note + " | " + tag).strip(" |")
                            if current_note else tag
                        )
                    break
            n_dupes += 1
        else:
            wd["source_video"] = video_title
            wd["source_link"]  = video_url
            wd["added_date"]   = today
            wd["note"]         = ""
            truly_new.append(wd)
            n_new += 1

    # New words at TOP, then existing (most recent first)
    combined = truly_new + existing_words
    return combined, n_new, n_dupes


# ─── Sheet Builders ────────────────────────────────────────────────────────────

COL_WIDTHS = {
    1: 5,   2: 10,  3: 22,  4: 22,  5: 16,  6: 14,
    7: 16,  8: 14,  9: 14,  10: 18, 11: 48, 12: 10,
    13: 10, 14: 38, 15: 4,  16: 7,  17: 7,  18: 7,
    19: 7,  20: 7,  21: 7,  22: 7,  23: 10, 24: 8,
    25: 8,  26: 18, 27: 30, 28: 38, 29: 14, 30: 40,
}

VOCAB_HEADERS = [
    "STT", "🔊 NGHE TỪ", "ENGLISH", "TIẾNG VIỆT",
    "✍️ GÕ CHỮ HÁN", "✅ KT HÁN", "🅰 GÕ PINYIN", "✅ KT PINYIN",
    "📖 ĐÁP ÁN HÁN", "🔤 PINYIN", "📝 VÍ DỤ",
    "🔊 NGHE VÍ DỤ", "🎬 YouGlish", "🔍 CHIẾT TỰ",
    "", "DAY 1", "DAY 2", "DAY 3", "DAY 4", "DAY 5", "DAY 6", "DAY 7",
    "NGÀY ÔN", "🔀 XÀO", "📊 TẦN SỐ", "📚 CẤP ĐỘ",
    "🎬 NGUỒN VIDEO",   # col 27 — video title
    "🔗 LINK VIDEO",    # col 28 — YouTube URL
    "📅 NGÀY THÊM",     # col 29 — date added
    "⚠️ GHI CHÚ",       # col 30 — duplicate / extra notes
]

# Column indices for reading back from a master xlsx (0-based, values_only)
MASTER_COL = {
    "stt": 0, "english": 2, "vietnamese": 3,
    "hanzi": 8, "pinyin": 9, "example": 10, "chiettu": 13,
    "freq": 24, "level": 25,
    "source_video": 26, "source_link": 27, "added_date": 28, "note": 29,
}


def build_info_sheet(ws, video_id, video_title, video_url,
                     in_vocab, new_words, lang_used):
    """Build the summary / info sheet."""
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 62
    ws.column_dimensions['C'].width = 22

    # Banner
    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value     = "🎬  HA — YouTube Vocabulary Extractor"
    c.font      = Font(name="Calibri", bold=True, color=HEADER_LIGHT, size=16)
    c.fill      = PatternFill(fill_type="solid", fgColor=VIDEO_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 44

    # ── Video metadata rows ──
    meta = [
        ("🎬 Tên Video",      video_title, ""),
        ("🔗 Link Video",     video_url,   ""),
        ("🆔 Video ID",       video_id,    ""),
        ("🌐 Ngôn ngữ phụ đề", lang_used,  ""),
        ("📅 Ngày xử lý",
         datetime.now().strftime("%Y-%m-%d  %H:%M"), ""),
        ("", "", ""),
    ]

    # ── Stats ──
    level_counts = Counter(w["level"] for w in in_vocab)
    total = len(in_vocab) + len(new_words)
    stats = [
        ("📊 THỐNG KÊ", "", ""),
        ("Tổng từ duy nhất", str(total), "từ"),
        ("✅ Đã có trong Vocab", str(len(in_vocab)),
         f"{100*len(in_vocab)//total if total else 0}%"),
        ("🆕 Từ mới chưa học", str(len(new_words)),
         f"{100*len(new_words)//total if total else 0}%"),
        ("", "", ""),
        ("📚 PHÂN BỔ CẤP ĐỘ", "", ""),
    ]
    for lvl in VOCAB_SHEETS:
        cnt = level_counts.get(lvl, 0)
        if cnt:
            stats.append((f"    {lvl}", str(cnt), "từ"))
    if new_words:
        stats.append(("    🆕 Mới", str(len(new_words)), "từ"))

    # ── Top words ──
    top = sorted(in_vocab + new_words, key=lambda x: -x["freq"])[:20]
    top_rows = [
        ("", "", ""),
        ("🔝 TOP TỪ XUẤT HIỆN NHIỀU NHẤT", "", ""),
    ]
    for i, w in enumerate(top, 1):
        meaning = w.get("english") or w.get("vietnamese") or "(chưa có)"
        top_rows.append((
            f"  {i:2d}.  {w['hanzi']}  ({w.get('pinyin', '')})",
            meaning,
            f"{w['freq']} lần  |  {w['level']}",
        ))

    all_rows = meta + stats + top_rows
    SECTION_HEADERS = {"📊 THỐNG KÊ", "📚 PHÂN BỔ CẤP ĐỘ", "🔝 TOP TỪ XUẤT HIỆN NHIỀU NHẤT"}

    for i, (label, val, note) in enumerate(all_rows, start=2):
        ws.row_dimensions[i].height = 18
        if not label and not val:
            continue
        if label in SECTION_HEADERS:
            ws.merge_cells(f"A{i}:C{i}")
            c = ws.cell(row=i, column=1, value=label)
            c.font      = Font(name="Calibri", bold=True, color=HEADER_LIGHT, size=11)
            c.fill      = PatternFill(fill_type="solid", fgColor="37474F")
            c.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[i].height = 22
            continue

        lc = ws.cell(row=i, column=1, value=label)
        lc.font      = Font(name="Calibri", bold=True, size=10)
        lc.alignment = Alignment(vertical="center", indent=1)

        if val.startswith("http"):
            vc2 = ws.cell(row=i, column=2, value=val)
            vc2.hyperlink = val
            vc2.font      = Font(name="Calibri", color=ACCENT_BLUE,
                                 underline="single", size=10)
        else:
            vc2 = ws.cell(row=i, column=2, value=val)
            vc2.font      = Font(name="Calibri", size=10)
        vc2.alignment = Alignment(vertical="center", wrap_text=True)

        if note:
            nc = ws.cell(row=i, column=3, value=note)
            nc.font      = Font(name="Calibri", size=9, italic=True, color="757575")
            nc.alignment = Alignment(vertical="center")

        if i % 2 == 0:
            for col in (1, 2, 3):
                c2 = ws.cell(row=i, column=col)
                if c2.fill.fill_type in (None, "none"):
                    c2.fill = PatternFill(fill_type="solid", fgColor="FAFAFA")


def build_vocab_sheet(ws, words: list, video_title: str, video_url: str):
    """Build a vocabulary sheet matching the Vocab (HSK tab) format."""
    _set_col_widths(ws, COL_WIDTHS)
    last_col = len(VOCAB_HEADERS)

    # Row 1 — title banner
    ws.merge_cells(f'A1:{get_column_letter(last_col)}1')
    c = ws['A1']
    c.value     = f"📖  HA  —  {video_title}"
    c.font      = Font(name="Calibri", bold=True, color=HEADER_LIGHT, size=13)
    c.fill      = PatternFill(fill_type="solid", fgColor=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Row 2 — video link
    ws.merge_cells(f'A2:{get_column_letter(last_col)}2')
    c = ws['A2']
    c.value     = f"🔗  {video_url}"
    c.hyperlink = video_url
    c.font      = Font(name="Calibri", color="90CAF9", underline="single", size=10)
    c.fill      = PatternFill(fill_type="solid", fgColor=VIDEO_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Row 3 — column headers
    for col, hdr in enumerate(VOCAB_HEADERS, start=1):
        _hdr(ws.cell(row=3, column=col, value=hdr))
    ws.row_dimensions[3].height = 36
    ws.freeze_panes = "C4"

    # Auto-filter on header row
    ws.auto_filter.ref = f"A3:{get_column_letter(last_col)}{3 + len(words)}"

    # Data rows
    for idx, wd in enumerate(words, start=1):
        r        = idx + 3
        hanzi    = wd["hanzi"]
        level    = wd.get("level", "🆕 Mới")
        freq     = wd.get("freq", 0)
        bg       = LEVEL_BG.get(level, LEVEL_BG["🆕 Mới"])
        example  = wd.get("example", "")

        ws.row_dimensions[r].height = 55 if example else 22

        # 1  STT
        _dat(ws.cell(r, 1, idx), bg=bg, align="center")

        # 2  Audio link
        _link(ws, r, 2, audio_url(hanzi), "🔊 Nghe", bg=bg)

        # 3  English
        _dat(ws.cell(r, 3, wd.get("english", "")), bg=bg)

        # 4  Vietnamese
        _dat(ws.cell(r, 4, wd.get("vietnamese", "")), bg=bg)

        # 5  Input Hán (practice — left blank)
        _dat(ws.cell(r, 5, ""), bg="FFFFFF")

        # 6  Check Hán
        c6 = ws.cell(r, 6, "⬜ Chưa làm")
        _dat(c6, bg="F5F5F5", align="center")
        c6.font = Font(name="Calibri", color="9E9E9E", size=9)

        # 7  Input Pinyin (practice — left blank)
        _dat(ws.cell(r, 7, ""), bg="FFFFFF")

        # 8  Check Pinyin
        c8 = ws.cell(r, 8, "⬜ Chưa làm")
        _dat(c8, bg="F5F5F5", align="center")
        c8.font = Font(name="Calibri", color="9E9E9E", size=9)

        # 9  Answer Hán
        c9 = ws.cell(r, 9, hanzi)
        _dat(c9, bg=bg, bold=True)
        c9.font = Font(name="Calibri", bold=True, size=14)

        # 10 Pinyin
        _dat(ws.cell(r, 10, wd.get("pinyin", "")), bg=bg)

        # 11 Example sentence
        c11 = ws.cell(r, 11, example)
        _dat(c11, bg=bg, wrap=True)
        c11.font = Font(name="Calibri", size=9)

        # 12 Audio example
        if example:
            ex_first_line = example.split('\n')[0]
            _link(ws, r, 12, audio_url(ex_first_line), "🔊 Nghe", bg=bg)
        else:
            _dat(ws.cell(r, 12, ""), bg=bg)

        # 13 YouGlish
        _link(ws, r, 13, youglish_url(hanzi), "🎬 Xem", bg=bg)

        # 14 Chiết tự
        c14 = ws.cell(r, 14, wd.get("chiettu", ""))
        _dat(c14, bg=bg, wrap=True)
        c14.font = Font(name="Calibri", size=9, color="4A4A4A")

        # 15 blank
        _dat(ws.cell(r, 15, ""), bg=bg)

        # 16–22 DAY 1–7
        for d in range(7):
            _dat(ws.cell(r, 16 + d, ""), bg="FFFFF0", align="center")

        # 23 NGÀY ÔN
        _dat(ws.cell(r, 23, 0), bg=bg, align="center")

        # 24 🔀 XÀO (random sort key)
        c24 = ws.cell(r, 24, round(random.random(), 6))
        _dat(c24, bg=bg, align="center")
        c24.font = Font(name="Calibri", size=8, color="BDBDBD")

        # 25 📊 Frequency
        freq_bg = "FFE082" if freq >= 10 else ("FFF9C4" if freq >= 5 else bg)
        c25 = ws.cell(r, 25, freq)
        _dat(c25, bg=freq_bg, align="center", bold=(freq >= 5))

        # 26 📚 Level
        c26 = ws.cell(r, 26, level)
        _dat(c26, bg=bg, align="center")
        c26.font = Font(name="Calibri", size=9, bold=True)

        # 27 🎬 Source Video
        src_video = wd.get("source_video", "")
        src_link  = wd.get("source_link",  "")
        if src_link:
            _link(ws, r, 27, src_link, src_video or src_link, bg="EDE7F6", fg="4527A0")
        else:
            c27 = ws.cell(r, 27, src_video)
            _dat(c27, bg="EDE7F6")
            c27.font = Font(name="Calibri", size=9)

        # 28 🔗 Link Video (plain URL for easy copy)
        c28 = ws.cell(r, 28, src_link)
        _dat(c28, bg="EDE7F6")
        c28.font = Font(name="Calibri", size=9, color=ACCENT_BLUE)

        # 29 📅 Date Added
        c29 = ws.cell(r, 29, wd.get("added_date", ""))
        _dat(c29, bg="EDE7F6", align="center")
        c29.font = Font(name="Calibri", size=9, color="757575")

        # 30 ⚠️ GHI CHÚ (duplicate / extra notes)
        note_val = wd.get("note", "")
        c30 = ws.cell(r, 30, note_val)
        note_bg = "FFF3E0" if note_val else "EDE7F6"
        _dat(c30, bg=note_bg, wrap=True)
        c30.font = Font(name="Calibri", size=9,
                        color="E65100" if note_val else "9E9E9E")


# ─── Main Output Builder ───────────────────────────────────────────────────────

def create_xlsx(
    video_id: str,
    video_title: str,
    video_url: str,
    lang_used: str,
    in_vocab: list,
    new_words: list,
    output_dir: str = ".",
    master_path: str = None,
) -> tuple[str, int, int]:
    """
    Assemble workbook and save.

    If master_path is given:
      - Load existing master data
      - Merge: new words on TOP, duplicates get notes updated
      - Save back to master_path (overwrite with full updated data)
      - Also save a dated per-video copy in output_dir

    Returns (output_file_path, n_new, n_duplicates).
    """
    today     = datetime.now().strftime("%Y-%m-%d")
    all_words = in_vocab + new_words

    # Stamp source info for non-master (single-video) mode
    for wd in all_words:
        if "source_video" not in wd:
            wd["source_video"] = video_title
            wd["source_link"]  = video_url
            wd["added_date"]   = today
            wd["note"]         = ""

    n_new   = len(all_words)
    n_dupes = 0

    if master_path:
        print(f"📂 Loading master file: {master_path}")
        existing_hanzi, existing_words = load_master_data(master_path)
        all_words, n_new, n_dupes = merge_with_master(
            all_words, existing_hanzi, existing_words, video_title, video_url
        )
        # Re-partition for info sheet stats
        in_vocab_merged  = [w for w in all_words if w.get("level", "") not in ("", "🆕 Mới")]
        new_words_merged = [w for w in all_words if w.get("level", "") in ("", "🆕 Mới")]
    else:
        in_vocab_merged  = in_vocab
        new_words_merged = new_words

    def _build_wb(words_list, iv, nw, title_suffix=""):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws_info = wb.create_sheet("📹 Video")
        build_info_sheet(ws_info, video_id, video_title, video_url,
                         iv, nw, lang_used)
        ws_all = wb.create_sheet("📚 Từ Vựng HA")
        build_vocab_sheet(ws_all, words_list, video_title + title_suffix, video_url)
        iv_sheet = [w for w in words_list if w.get("level", "") not in ("", "🆕 Mới")]
        nw_sheet = [w for w in words_list if w.get("level", "") in ("", "🆕 Mới")]
        if iv_sheet:
            ws_k = wb.create_sheet("✅ Đã Học")
            build_vocab_sheet(ws_k, iv_sheet, video_title + title_suffix, video_url)
        if nw_sheet:
            ws_n = wb.create_sheet("🆕 Từ Mới")
            build_vocab_sheet(ws_n, nw_sheet, video_title + title_suffix, video_url)
        return wb

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M')

    # Per-video dated file (always created)
    wb_video = _build_wb(
        all_words if not master_path else
        [w for w in all_words if w.get("source_video", "") == video_title],
        in_vocab, new_words,
    )
    fname      = f"HA_{safe_filename(video_title)}_{stamp}.xlsx"
    video_path = Path(output_dir) / fname
    wb_video.save(str(video_path))

    # Master accumulation file
    if master_path:
        wb_master = _build_wb(all_words, in_vocab_merged, new_words_merged,
                               title_suffix=" [MASTER]")
        wb_master.save(master_path)
        print(f"   ✅ Master updated: {master_path}")

    return str(video_path), n_new, n_dupes


# ─── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        prog="HA",
        description="YouTube Vocabulary Extractor — map video words to Vocab (HSK) format",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python HA.py "https://www.youtube.com/watch?v=dQw4w9WgXcQ"
  python HA.py "https://youtu.be/VIDEO_ID" --vocab "./TuVung_Trung-Anh-Viet.xlsx"
  python HA.py VIDEO_ID --min-freq 2 --min-chars 2 --output "./output"
  python HA.py VIDEO_ID --no-stopwords --lang zh-TW
        """,
    )
    parser.add_argument("url",
        help="YouTube URL, youtu.be link, or bare video ID")
    parser.add_argument("--vocab", "-v", default=None,
        help="Path to Vocab xlsx (auto-detected if omitted)")
    parser.add_argument("--output", "-o", default=".",
        help="Output folder for the generated xlsx (default: current dir)")
    parser.add_argument("--min-freq", "-f", type=int, default=1,
        help="Minimum word frequency to include (default: 1)")
    parser.add_argument("--min-chars", "-c", type=int, default=1,
        help="Minimum Chinese character count per word (default: 1)")
    parser.add_argument("--no-stopwords", action="store_true",
        help="Filter out common function words / particles")
    parser.add_argument("--lang", default=None,
        help="Force transcript language code, e.g. zh-Hans, zh-TW, en")
    parser.add_argument("--top", "-t", type=int, default=100,
        help="Keep only top N words by frequency (default: 100)")
    parser.add_argument("--master", "-m", default=None,
        help="Path to master accumulation xlsx.  New words go to TOP; "
             "duplicates get a note instead of being re-added.")

    args = parser.parse_args()

    sep = "=" * 62
    print(f"\n{sep}")
    print("🎬  HA — YouTube Vocabulary Extractor")
    print(sep)

    # ── 1. Parse URL ──
    try:
        video_id  = extract_video_id(args.url)
        video_url = f"https://www.youtube.com/watch?v={video_id}"
        print(f"📹 Video ID : {video_id}")
    except ValueError as e:
        print(f"❌  {e}")
        sys.exit(1)

    # ── 2. Video title ──
    print("📡 Fetching video title ...")
    video_title = get_video_title(video_id)
    print(f"   📋 {video_title}")

    # ── 3. Transcript ──
    print("📝 Fetching transcript ...")
    try:
        segments, lang_used = fetch_transcript(video_id, force_lang=args.lang)
        print(f"   Segments: {len(segments):,}")
    except RuntimeError as e:
        print(f"❌  {e}")
        sys.exit(1)

    # ── 4. Segment ──
    print("✂️  Segmenting with jieba ...")
    raw_text = segments_to_text(segments)
    total_zh = sum(1 for c in raw_text if is_chinese_char(c))
    print(f"   Chinese characters: {total_zh:,}")

    word_freq = segment_chinese(raw_text, min_chars=args.min_chars)

    if args.min_freq > 1:
        word_freq = Counter({w: f for w, f in word_freq.items()
                             if f >= args.min_freq})
    print(f"   Unique words: {len(word_freq):,}")

    # ── 5. Vocab DB ──
    vocab_path = args.vocab
    if not vocab_path:
        vocab_path = auto_detect_vocab(Path(__file__).parent)
        if vocab_path:
            print(f"🔍 Auto-detected Vocab: {vocab_path}")

    print("📚 Loading Vocab database ...")
    vocab_db = load_vocab(vocab_path)

    # ── 6. Map ──
    print("🗺️  Mapping words ...")
    in_vocab, new_words = map_to_vocab(
        word_freq, vocab_db, remove_stopwords=args.no_stopwords
    )

    lvl_counts = Counter(w["level"] for w in in_vocab)
    print(f"   ✅ In Vocab  : {len(in_vocab):,}")
    for lvl in VOCAB_SHEETS:
        if lvl in lvl_counts:
            print(f"      {lvl}: {lvl_counts[lvl]}")
    print(f"   🆕 New words : {len(new_words):,}")

    # ── Apply top-N filter ──
    all_words = in_vocab + new_words
    all_words_sorted = sorted(all_words, key=lambda x: -x["freq"])
    top_words = all_words_sorted[:args.top]
    in_vocab  = [w for w in top_words if w.get("level", "") not in ("", "🆕 Mới")]
    new_words = [w for w in top_words if w.get("level", "") in ("", "🆕 Mới")]
    print(f"\n🏆 Top {args.top} by frequency: {len(top_words)} words")

    # ── 7. Export ──
    print("\n📊 Building Excel file ...")
    out_path, n_new, n_dupes = create_xlsx(
        video_id=video_id,
        video_title=video_title,
        video_url=video_url,
        lang_used=lang_used,
        in_vocab=in_vocab,
        new_words=new_words,
        output_dir=args.output,
        master_path=args.master,
    )

    master_note = ""
    if args.master:
        master_note = (
            f"\n   🆕 New words added   : {n_new}"
            f"\n   🔁 Duplicates noted  : {n_dupes}"
            f"\n   📂 Master updated    : {args.master}"
        )

    print(f"\n✅  Saved to:")
    print(f"   {out_path}{master_note}")
    print(f"""
💡  TIPS
   • Open '📚 Từ Vựng HA'  — full study sheet (same format as Vocab)
   • Open '✅ Đã Học'       — review words you already know
   • Open '🆕 Từ Mới'       — add these to your main Vocab file
   • Sort column 📊 TẦN SỐ  — study the most frequent words first
   • Sort column 🔀 XÀO      — shuffle for random drilling
   • Column ⚠️ GHI CHÚ       — shows which other videos contain the same word
   • Column 🎬 NGUỒN VIDEO   — click to open the source video
   • Hide columns E,F,G,H   — self-test mode (no answers visible)
   • Use --master HA_master.xlsx to accumulate words across all videos
{sep}
""")


if __name__ == "__main__":
    main()
