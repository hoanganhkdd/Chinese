#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HA_video.py — YouTube ➜ Từ vựng ➜ ghi thẳng vào HSK1_Tu_Vung_500.xlsx
=======================================================================
QUY TRÌNH
  1. Xem 1 video → dán link vào ô B2 sheet "⚙️ NHẬP VIDEO" trong file Excel
  2. Script lấy chữ Hán từ video → jieba tách từ
  3. Đối chiếu cột C (🀄 HÁN TỰ) → BỎ từ đã có, CHỈ giữ từ MỚI
  4. Tự điền đủ 23 cột (pinyin, nghĩa Việt, audio, chiết tự, câu ví dụ…)
  5. Cột U (📂 CHỦ ĐỀ) = tên video + link bấm được
  → Từ mới nối xuống CUỐI file, dữ liệu cũ giữ nguyên 100%

v2.0 — 6 TẦNG LẤY CHỮ HÁN (tự động thử lần lượt)
  1️⃣  Phụ đề CC tiếng Trung (thủ công)
  2️⃣  Phụ đề CC tự động (ASR) tiếng Trung
  3️⃣  Bất kỳ track phụ đề nào CÓ CHỨA chữ Hán  ← phụ đề song ngữ Việt–Trung
  4️⃣  Phần MÔ TẢ video                          ← kênh Việt hay dán hội thoại ở đây
  5️⃣  Sheet "📋 DÁN PHỤ ĐỀ" trong Excel          ← bạn tự dán, luôn chạy được
  6️⃣  OCR chữ nung cứng trên hình (--ocr)        ← video slideshow, hard-sub

CÁCH DÙNG
  double-click HA_video.bat
  python HA_video.py
  python HA_video.py "https://youtu.be/XXXX"
  python HA_video.py --check "https://youtu.be/XXXX"     ← chẩn đoán video
  python HA_video.py --ocr   "https://youtu.be/XXXX"     ← video chữ nung cứng
  python HA_video.py --text phude.txt --title "Tên video"

CÀI THƯ VIỆN
  pip install youtube-transcript-api jieba pypinyin openpyxl yt-dlp zhconv
  (OCR cần thêm: pip install pytesseract pillow  +  Tesseract-OCR + gói chi_sim
   + ffmpeg)

Version 2.0 · 2026-08-02
"""

import argparse
import glob
import html
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import urllib.parse
import urllib.request
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path

# ── Ép UTF-8 cho CMD Windows (tránh UnicodeEncodeError khi in chữ Hán) ─────────
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# ─────────────────────────── Kiểm tra thư viện ────────────────────────────────
MISSING = []
try:
    import jieba
    jieba.setLogLevel(60)
except ImportError:
    MISSING.append("jieba")
try:
    from pypinyin import pinyin as _pinyin, Style
except ImportError:
    MISSING.append("pypinyin")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    MISSING.append("openpyxl")

if MISSING:
    sys.exit("❌ Thiếu thư viện. Chạy lệnh:\n   pip install " + " ".join(MISSING))

# ─────────────────────────────── Hằng số ──────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = SCRIPT_DIR / "HSK1_Tu_Vung_500.xlsx"
CACHE_FILE = SCRIPT_DIR / "HA_char_cache.json"

SHEET_VOCAB = "📚 Từ Vựng HSK"
SHEET_INPUT = "⚙️ NHẬP VIDEO"
SHEET_PASTE = "📋 DÁN PHỤ ĐỀ"

HEADER_ROW = 3
FIRST_DATA_ROW = 4
N_COLS = 23  # A..W

C_STT, C_LEVEL, C_HAN, C_PIN, C_VIET = 1, 2, 3, 4, 5
C_TRY1, C_TRY2, C_CHECK, C_AUDIO, C_YOUG = 6, 7, 8, 9, 10
C_DATE, C_NOTE, C_CHIET, C_TOPIC, C_EX, C_EXPIN = 18, 19, 20, 21, 22, 23

ROW_HEIGHT = 69.75

CELL_LINK = "B2"
CELL_TOPN = "B3"
CELL_MINCHARS = "B4"
CELL_STOPWORDS = "B5"
CELL_LANG = "B6"
CELL_STATUS = "B7"
CELL_SOURCE = "B8"
HISTORY_START_ROW = 11
PASTE_START_ROW = 4

STOPWORDS = set("""
的 了 是 在 我 你 他 她 它 们 这 那 有 和 就 不 人 都 一 一个 上 也 很 到 说 要 去
会 着 没有 看 好 自己 这个 那个 什么 怎么 可以 但是 因为 所以 如果 还是 已经 还有
我们 你们 他们 她们 一样 这样 那样 时候 现在 知道 觉得 一下 一点 一些 大家 真的
其实 就是 可能 应该 然后 而且 或者 不是 没 也是 那么 一直 比较 非常 特别 只是
个 吧 吗 呢 啊 呀 哦 嗯 哈 嘛 呗 咯 之 与 及 于 里 中 下 前 后 来 过 得 把 被 给
对 从 向 让 使 做 想 用 多 少 大 小 新 老 高 低 快 慢
""".split())

# Chữ Hán hay xuất hiện trong logo/watermark của kênh dạy học → bỏ
NOISE_WORDS = {"字幕", "翻译", "订阅", "点赞", "关注", "频道", "视频", "本期", "下期"}

CJK_PUNCT = "，。！？、；：""''（）《》…—～·"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/122.0 Safari/537.36")


def log(msg=""):
    print(msg, flush=True)


def safe_save(wb, path, quiet=False):
    try:
        wb.save(path)
        return True
    except PermissionError:
        if not quiet:
            log("\n❌ KHÔNG GHI ĐƯỢC — file đang mở trong Excel.")
            log("   ➜ Đóng Excel (nhớ Lưu trước), rồi chạy lại HA_video.bat")
        return False


# ───────────────────────────── Tiện ích chữ Hán ───────────────────────────────
def is_cjk(c):
    return "一" <= c <= "鿿"


def has_cjk(s):
    return any(is_cjk(c) for c in s)


def all_cjk(s):
    return bool(s) and all(is_cjk(c) for c in s)


def cjk_ratio(s):
    if not s:
        return 0.0
    return sum(1 for c in s if is_cjk(c)) / len(s)


_T2S = None


def to_simplified(s):
    """Phồn thể ➜ giản thể (nếu có zhconv / opencc; không có thì giữ nguyên)."""
    global _T2S
    if _T2S is None:
        try:
            from zhconv import convert as _cv
            _T2S = lambda t: _cv(t, "zh-cn")
        except Exception:
            try:
                from opencc import OpenCC
                _T2S = OpenCC("t2s").convert
            except Exception:
                _T2S = lambda t: t
    try:
        return _T2S(s)
    except Exception:
        return s


def to_pinyin(word):
    """Pinyin của TỪ — viết liền (hòumiàn, cāngkù) đúng kiểu file gốc."""
    try:
        return "".join(x[0] for x in _pinyin(word, style=Style.TONE)).strip()
    except Exception:
        return ""


def sentence_pinyin(s):
    """Pinyin của CÂU — cách nhau bởi dấu cách, dấu câu dính liền."""
    try:
        out = " ".join(x[0] for x in _pinyin(s, style=Style.TONE))
        out = re.sub(r"\s+([，。！？、；：）,\.\!\?\)])", r"\1", out)
        out = re.sub(r"([（\(])\s+", r"\1", out)
        return re.sub(r"\s{2,}", " ", out).strip()
    except Exception:
        return ""


def audio_url(word):
    q = urllib.parse.quote(word)
    return f"https://translate.google.com/?sl=zh-CN&tl=vi&text={q}&op=translate"


def youglish_url(word):
    return f"https://youglish.com/pronounce/{urllib.parse.quote(word)}/chinese"


# ──────────────────────────── Dịch (Google free) ──────────────────────────────
_trans_cache = {}


def gtranslate(texts, sl="zh-CN", tl="vi", pause=0.35):
    out, todo = {}, []
    for t in texts:
        t = (t or "").strip()
        if not t:
            continue
        if t in _trans_cache:
            out[t] = _trans_cache[t]
        else:
            todo.append(t)

    BATCH = 15
    for i in range(0, len(todo), BATCH):
        chunk = todo[i:i + BATCH]
        joined = "\n".join(chunk)
        try:
            url = ("https://translate.googleapis.com/translate_a/single"
                   f"?client=gtx&sl={sl}&tl={tl}&dt=t&q={urllib.parse.quote(joined)}")
            req = urllib.request.Request(url, headers={"User-Agent": UA})
            with urllib.request.urlopen(req, timeout=20) as r:
                data = json.loads(r.read().decode("utf-8", "ignore"))
            merged = "".join(seg[0] for seg in data[0] if seg and seg[0])
            lines = [l.strip() for l in merged.split("\n")]
            if len([l for l in lines if l]) != len(chunk):
                raise ValueError("lệch dòng")
            for src, dst in zip(chunk, lines):
                dst = dst.strip().rstrip(".")
                _trans_cache[src] = dst
                out[src] = dst
        except Exception:
            for src in chunk:
                try:
                    url = ("https://translate.googleapis.com/translate_a/single"
                           f"?client=gtx&sl={sl}&tl={tl}&dt=t&q={urllib.parse.quote(src)}")
                    req = urllib.request.Request(url, headers={"User-Agent": UA})
                    with urllib.request.urlopen(req, timeout=15) as r:
                        data = json.loads(r.read().decode("utf-8", "ignore"))
                    dst = "".join(s[0] for s in data[0] if s and s[0]).strip().rstrip(".")
                except Exception:
                    dst = ""
                _trans_cache[src] = dst
                out[src] = dst
                time.sleep(pause)
        time.sleep(pause)
        log(f"   … đã dịch {min(i + BATCH, len(todo))}/{len(todo)}")
    return out


# ═════════════════════ LÀM SẠCH & CHUẨN HOÁ DÒNG CHỮ HÁN ══════════════════════
def chunk_line(s, maxlen=30):
    """Cắt câu quá dài (phụ đề ASR không có dấu câu) thành đoạn ngắn dễ đọc."""
    if len(s) <= maxlen:
        return [s]
    parts, buf = [], ""
    for ch in s:
        buf += ch
        if ch in "，。！？；、" and len(buf) >= 8:
            parts.append(buf); buf = ""
        elif len(buf) >= maxlen:
            parts.append(buf); buf = ""
    if buf:
        parts.append(buf)
    return [p.strip() for p in parts if p.strip()]


def clean_lines(raw_lines):
    """
    • Bỏ thẻ HTML, [音乐], [Âm nhạc]…
    • Tách RIÊNG phần chữ Hán ra khỏi dòng song ngữ Việt–Trung / có pinyin
    • Phồn thể ➜ giản thể
    • Bỏ dòng trùng liên tiếp (phụ đề tự động hay lặp)
    • Cắt câu quá dài
    """
    out, prev = [], ""
    for ln in raw_lines:
        ln = html.unescape(str(ln))
        ln = re.sub(r"<[^>]+>", " ", ln)
        ln = re.sub(r"\[[^\]]{0,25}\]", " ", ln)
        ln = re.sub(r"\([^)]{0,25}\)", " ", ln)
        ln = ln.replace("​", " ").replace("\xa0", " ")
        if not has_cjk(ln):
            continue
        # giữ chữ Hán + dấu câu Trung, mọi thứ khác thành khoảng trắng ngăn cách
        masked = "".join(c if (is_cjk(c) or c in CJK_PUNCT) else " " for c in ln)
        for frag in masked.split(" "):
            frag = frag.strip(CJK_PUNCT + " ")
            if len(frag) < 2 or not has_cjk(frag):
                continue
            frag = to_simplified(frag)
            for piece in chunk_line(frag):
                if piece and piece != prev:
                    out.append(piece)
                    prev = piece
    return out


# ═══════════════════════════ NGUỒN LẤY CHỮ HÁN ════════════════════════════════
def extract_video_id(url):
    url = (url or "").strip().strip('"').strip("'")
    if not url:
        return None
    if re.fullmatch(r"[A-Za-z0-9_-]{11}", url):
        return url
    for p in (r"(?:youtube\.com/watch\?[^ ]*?v=)([A-Za-z0-9_-]{11})",
              r"(?:youtu\.be/)([A-Za-z0-9_-]{11})",
              r"(?:youtube\.com/(?:embed|v|shorts|live)/)([A-Za-z0-9_-]{11})"):
        m = re.search(p, url)
        if m:
            return m.group(1)
    return None


def get_video_title(vid):
    try:
        url = (f"https://www.youtube.com/oembed?url="
               f"https://www.youtube.com/watch?v={vid}&format=json")
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=15) as r:
            d = json.loads(r.read().decode("utf-8", "ignore"))
            return d.get("title") or f"Video {vid}"
    except Exception:
        return f"Video {vid}"


def _vtt_to_lines(raw):
    lines = []
    for ln in raw.splitlines():
        ln = ln.strip()
        if (not ln or "-->" in ln
                or ln.startswith(("WEBVTT", "Kind:", "Language:", "NOTE", "STYLE"))
                or re.fullmatch(r"\d+", ln)):
            continue
        lines.append(ln)
    return lines


def parse_sub_payload(raw, ext=""):
    """Hiểu được json3 / srv3 / vtt / ttml / srv1(xml) / txt."""
    raw = raw.strip()
    if raw.startswith("{"):
        try:
            data = json.loads(raw)
            lines = []
            for ev in data.get("events", []):
                t = "".join(s.get("utf8", "") for s in (ev.get("segs") or []))
                t = t.replace("\n", " ").strip()
                if t:
                    lines.append(t)
            if lines:
                return lines
        except Exception:
            pass
    if raw.startswith("WEBVTT") or "-->" in raw:
        return _vtt_to_lines(raw)
    if "<text" in raw or "<p " in raw:
        return re.findall(r"<(?:text|p)[^>]*>(.*?)</(?:text|p)>", raw, re.S)
    return raw.splitlines()


def _http_get(url, timeout=25):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read().decode("utf-8", "ignore")


# ── Nguồn A: youtube-transcript-api ──────────────────────────────────────────
def src_transcript_api(vid, langs):
    try:
        from youtube_transcript_api import YouTubeTranscriptApi as API
    except ImportError:
        return [], None
    try:
        if hasattr(API, "fetch") and not hasattr(API, "get_transcript"):
            api = API()
            try:
                fetched = api.fetch(vid, languages=langs)
                return [s.text for s in fetched], "CC tiếng Trung"
            except Exception:
                tl = api.list(vid)
                for tr in tl:
                    got = tr.fetch()
                    lines = [s.text for s in got]
                    if any(has_cjk(l) for l in lines):
                        kind = "tự động" if getattr(tr, "is_generated", False) else "thủ công"
                        return lines, f"CC [{tr.language_code}] ({kind})"
        else:
            try:
                segs = API.get_transcript(vid, languages=langs)
                return [s.get("text", "") for s in segs], "CC tiếng Trung"
            except Exception:
                tl = API.list_transcripts(vid)
                for tr in tl:
                    lines = [s.get("text", "") for s in tr.fetch()]
                    if any(has_cjk(l) for l in lines):
                        kind = "tự động" if getattr(tr, "is_generated", False) else "thủ công"
                        return lines, f"CC [{tr.language_code}] ({kind})"
    except Exception as e:
        log(f"   ⚠️ youtube-transcript-api: {type(e).__name__}")
    return [], None


# ── yt-dlp: metadata + mọi track phụ đề ──────────────────────────────────────
class _QuietLogger:
    def debug(self, m): pass
    def info(self, m): pass
    def warning(self, m): pass
    def error(self, m): pass


def ytdlp_info(vid):
    try:
        from yt_dlp import YoutubeDL
    except ImportError:
        log("   ⚠️ Chưa cài yt-dlp  (pip install yt-dlp)")
        return None
    opts = {"skip_download": True, "quiet": True, "no_warnings": True,
            "logger": _QuietLogger(),
            "writesubtitles": True, "writeautomaticsub": True,
            "extractor_args": {"youtube": {"player_client": ["web", "android"]}}}
    try:
        with YoutubeDL(opts) as ydl:
            return ydl.extract_info(f"https://www.youtube.com/watch?v={vid}", download=False)
    except Exception as e:
        log(f"   ⚠️ yt-dlp: {type(e).__name__} — {str(e)[:90]}")
        return None


def list_tracks(info):
    """[(lang, is_auto, url, ext)] — ưu tiên định dạng json3."""
    tracks = []
    for key, auto in (("subtitles", False), ("automatic_captions", True)):
        for lang, fmts in (info.get(key) or {}).items():
            best = None
            for f in fmts or []:
                if f.get("ext") in ("json3", "srv3", "srv1", "vtt", "ttml"):
                    if best is None or f["ext"] == "json3":
                        best = f
            if best and best.get("url"):
                tracks.append((lang, auto, best["url"], best.get("ext", "")))
    # sắp xếp: tiếng Trung thủ công → Trung tự động → còn lại
    def rank(t):
        lang, auto = t[0], t[1]
        zh = lang.lower().startswith("zh")
        return (0 if zh and not auto else 1 if zh else 2, 0 if not auto else 1, lang)
    return sorted(tracks, key=rank)


def src_ytdlp_tracks(info):
    for lang, auto, url, ext in list_tracks(info):
        try:
            raw = _http_get(url)
        except Exception:
            continue
        lines = parse_sub_payload(raw, ext)
        if any(has_cjk(l) for l in lines):
            kind = "tự động" if auto else "thủ công"
            return lines, f"CC [{lang}] ({kind})"
    return [], None


def src_description(info):
    desc = (info or {}).get("description") or ""
    if not has_cjk(desc):
        return [], None
    lines = [l for l in desc.splitlines() if has_cjk(l)]
    if len("".join(lines)) < 30:
        return [], None
    return lines, "MÔ TẢ video"


# ── Nguồn OCR: chữ nung cứng trên hình ───────────────────────────────────────
def src_ocr(vid, every=3, quality="360"):
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        log("   ⚠️ OCR cần:  pip install pytesseract pillow")
        return [], None
    if not shutil.which("ffmpeg"):
        log("   ⚠️ OCR cần ffmpeg  (https://www.gyan.dev/ffmpeg/builds/ → thêm vào PATH)")
        return [], None
    try:
        langs = pytesseract.get_languages(config="")
        if "chi_sim" not in langs:
            log("   ⚠️ Tesseract chưa có gói tiếng Trung 'chi_sim'.")
            log("      Cài Tesseract (UB-Mannheim build) và tick 'Chinese Simplified'.")
            return [], None
    except Exception:
        log("   ⚠️ Chưa cài Tesseract-OCR hoặc chưa có trong PATH.")
        return [], None

    tmp = tempfile.mkdtemp(prefix="haocr_")
    try:
        log("   ⏬ Đang tải video (bản nhẹ) để OCR…")
        from yt_dlp import YoutubeDL
        vpath = os.path.join(tmp, "v.mp4")
        with YoutubeDL({"quiet": True, "no_warnings": True, "outtmpl": vpath,
                        "logger": _QuietLogger(),
                        "format": f"best[height<={quality}]/worst"}) as ydl:
            ydl.download([f"https://www.youtube.com/watch?v={vid}"])
        if not os.path.exists(vpath):
            cands = glob.glob(os.path.join(tmp, "v.*"))
            if not cands:
                return [], None
            vpath = cands[0]

        fdir = os.path.join(tmp, "f")
        os.makedirs(fdir, exist_ok=True)
        log(f"   🖼️ Đang cắt hình mỗi {every}s…")
        subprocess.run(["ffmpeg", "-i", vpath, "-vf", f"fps=1/{every}",
                        "-q:v", "3", os.path.join(fdir, "%05d.jpg")],
                       capture_output=True, timeout=900)
        frames = sorted(glob.glob(os.path.join(fdir, "*.jpg")))
        if not frames:
            return [], None
        log(f"   🔎 OCR {len(frames)} khung hình (có thể mất vài phút)…")
        lines, prev = [], ""
        for i, f in enumerate(frames, 1):
            try:
                txt = pytesseract.image_to_string(Image.open(f), lang="chi_sim")
            except Exception:
                continue
            for ln in txt.splitlines():
                ln = re.sub(r"\s+", "", ln)
                if len(ln) >= 2 and cjk_ratio(ln) > 0.6 and ln != prev:
                    lines.append(ln)
                    prev = ln
            if i % 25 == 0:
                log(f"      … {i}/{len(frames)}")
        return lines, "OCR chữ trên hình"
    except Exception as e:
        log(f"   ⚠️ OCR lỗi: {type(e).__name__} — {str(e)[:90]}")
        return [], None
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def src_paste_sheet(wb):
    if SHEET_PASTE not in wb.sheetnames:
        return [], None
    ws = wb[SHEET_PASTE]
    lines = []
    for r in range(PASTE_START_ROW, ws.max_row + 1):
        for c in (1, 2):
            v = ws.cell(r, c).value
            if v and has_cjk(str(v)):
                lines.append(str(v))
    if not lines:
        return [], None
    return lines, "sheet 📋 DÁN PHỤ ĐỀ"


# ── Bộ điều phối: thử lần lượt 6 nguồn ───────────────────────────────────────
def gather_chinese(vid, wb, lang_pref="", want_ocr=False, mode="auto"):
    langs = [l.strip() for l in (lang_pref or "").split(",") if l.strip()]
    langs += ["zh-Hans", "zh-CN", "zh", "zh-Hant", "zh-TW", "zh-HK"]
    langs = list(OrderedDict.fromkeys(langs))
    info = None

    def ok(lines, src):
        clean = clean_lines(lines)
        return (clean, src) if len("".join(clean)) >= 30 else (None, None)

    order = []
    if mode == "dan":
        order = ["paste"]
    elif mode == "mota":
        order = ["desc", "paste"]
    elif mode == "ocr":
        order = ["ocr", "paste"]
    else:
        order = ["api", "ytdlp", "desc", "paste"] + (["ocr"] if want_ocr else [])

    for step in order:
        if step == "api":
            log("   ① Thử phụ đề CC tiếng Trung…")
            lines, src = src_transcript_api(vid, langs)
            c, s = ok(lines, src)
            if c:
                return c, s, info
        elif step == "ytdlp":
            log("   ② Thử mọi track phụ đề (kể cả song ngữ)…")
            info = info or ytdlp_info(vid)
            if info:
                lines, src = src_ytdlp_tracks(info)
                c, s = ok(lines, src)
                if c:
                    return c, s, info
        elif step == "desc":
            log("   ③ Thử phần mô tả video…")
            info = info or ytdlp_info(vid)
            if info:
                lines, src = src_description(info)
                c, s = ok(lines, src)
                if c:
                    return c, s, info
        elif step == "paste":
            lines, src = src_paste_sheet(wb)
            if lines:
                log("   ④ Dùng sheet 📋 DÁN PHỤ ĐỀ…")
                c, s = ok(lines, src)
                if c:
                    return c, s, info
        elif step == "ocr":
            log("   ⑤ OCR chữ nung cứng trên hình…")
            lines, src = src_ocr(vid)
            c, s = ok(lines, src)
            if c:
                return c, s, info
    return [], None, info


def diagnose(vid, wb):
    """--check : báo cáo video có gì, nên dùng cách nào."""
    log("\n🩺 CHẨN ĐOÁN VIDEO")
    log("─" * 62)
    title = get_video_title(vid)
    log(f"🎬 {title}")
    info = ytdlp_info(vid)
    if not info:
        log("❌ Không đọc được metadata bằng yt-dlp.")
        lines, src = src_transcript_api(vid, ["zh-Hans", "zh-CN", "zh", "zh-Hant", "zh-TW"])
        if lines:
            log(f"✅ Nhưng LẤY ĐƯỢC phụ đề qua youtube-transcript-api ({src}) — "
                "chạy HA_video.bat bình thường được.")
        else:
            log("   Nguyên nhân thường gặp: mạng/proxy chặn YouTube, hoặc yt-dlp cũ.")
            log("   ➜ Thử:  python -m pip install -U yt-dlp")
            log("   ➜ Hoặc dùng sheet '📋 DÁN PHỤ ĐỀ' (luôn chạy được).")
        return
    dur = info.get("duration") or 0
    log(f"⏱️  Thời lượng: {dur // 60}:{dur % 60:02d}   ·   Kênh: {info.get('uploader', '?')}")

    tracks = list_tracks(info)
    if tracks:
        log(f"\n📝 Có {len(tracks)} track phụ đề:")
        for lang, auto, url, ext in tracks[:20]:
            mark = "🤖 tự động" if auto else "✍️ thủ công"
            zh = " ⬅️ TIẾNG TRUNG" if lang.lower().startswith("zh") else ""
            log(f"     • {lang:<10} {mark} ({ext}){zh}")
    else:
        log("\n📝 KHÔNG có track phụ đề nào.")

    zh_tracks = [t for t in tracks if t[0].lower().startswith("zh")]
    desc = info.get("description") or ""
    n_han = sum(1 for c in desc if is_cjk(c))
    log(f"\n📄 Mô tả video: {len(desc)} ký tự, trong đó {n_han} chữ Hán"
        f"{' ✅ dùng được' if n_han >= 30 else ' ❌ không đủ'}")

    log("\n💡 KẾT LUẬN")
    if zh_tracks:
        log("   ✅ Có phụ đề tiếng Trung — chạy bình thường:  HA_video.bat")
    elif any(has_cjk(str(t[0])) for t in tracks) or tracks:
        log("   ⚠️ Không có track tiếng Trung riêng, nhưng script sẽ tự dò")
        log("      mọi track xem có chữ Hán không (phụ đề song ngữ).")
    if n_han >= 30:
        log("   ✅ Mô tả video có hội thoại chữ Hán — script sẽ tự lấy.")
    if not zh_tracks and n_han < 30:
        log("   ❗ Nhiều khả năng chữ Hán bị NUNG CỨNG vào hình. Hai cách:")
        log("      (a) Dễ nhất — mở sheet '📋 DÁN PHỤ ĐỀ', dán chữ Hán vào cột A")
        log("          (chép từ mô tả video, từ ảnh, hoặc gõ lại), rồi chạy lại.")
        log("      (b) Tự động — cài Tesseract + ffmpeg rồi chạy:")
        log("          python HA_video.py --ocr \"<link>\"")
    log("─" * 62)


# ───────────────────────── Tách từ & chọn từ mới ──────────────────────────────
def segment(lines, min_chars, skip_stop):
    freq, example = Counter(), {}
    for ln in lines:
        for w in jieba.cut(ln, cut_all=False):
            w = w.strip()
            if not all_cjk(w) or len(w) < min_chars:
                continue
            if skip_stop and w in STOPWORDS:
                continue
            if w in NOISE_WORDS:
                continue
            freq[w] += 1
            cand = ln.strip()
            if 4 <= len(cand) <= 40 and (w not in example or len(cand) < len(example[w])):
                example[w] = cand
    return freq, example


# ────────────────────── Cache chiết tự (học từ chính file) ────────────────────
def build_char_cache(ws, last_row):
    cache = {}
    for r in range(FIRST_DATA_ROW, last_row + 1):
        t = ws.cell(r, C_CHIET).value
        if not t or not isinstance(t, str):
            continue
        if "·" in t:
            for b in re.split(r"\n(?=[一-鿿]\s*\[)", t):
                m = re.match(r"^([一-鿿])\s*\[([^\]]*)\]\s*·\s*(.*?)"
                             r"(?:\n\s*↳\s*Chi[ếe]t t[ựu]:\s*(.*))?$", b.strip(), re.S)
                if m:
                    cache.setdefault(m.group(1), {
                        "hv": m.group(2).strip(),
                        "mean": m.group(3).strip().replace("\n", " "),
                        "chiet": (m.group(4) or "").strip().replace("\n", " "),
                    })
        else:
            for m in re.finditer(r"([一-鿿])\s*\(([^()]*(?:\([^()]*\)[^()]*)*)\)", t):
                parts = re.split(r"\s+[—–-]\s+", m.group(2), 1)
                cache.setdefault(m.group(1), {
                    "hv": "", "mean": parts[0].strip(),
                    "chiet": parts[1].strip() if len(parts) > 1 else "",
                })
    if CACHE_FILE.exists():
        try:
            for k, v in json.loads(CACHE_FILE.read_text(encoding="utf-8")).items():
                cache.setdefault(k, v)
        except Exception:
            pass
    return cache


def save_cache(cache):
    try:
        CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=0),
                              encoding="utf-8")
    except Exception:
        pass


def build_chiet_tu(word, cache):
    out = []
    for ch in word:
        if not is_cjk(ch):
            continue
        info = cache.get(ch)
        if info:
            hv = f" [{info['hv']}]" if info.get("hv") else ""
            line = f"{ch}{hv} · {info.get('mean', '')}".rstrip(" ·")
            if info.get("chiet"):
                line += f"\n    ↳ Chiết tự: {info['chiet']}"
        else:
            line = f"{ch} · (chưa có dữ liệu — bổ sung sau)"
        out.append(line)
    return "\n".join(out)


# ──────────────────────────── Sheet ⚙️ NHẬP VIDEO ─────────────────────────────
def ensure_input_sheet(wb):
    if SHEET_INPUT in wb.sheetnames:
        ws = wb[SHEET_INPUT]
        if not ws["A8"].value:  # nâng cấp từ v1 lên v2
            ws["A8"] = "🎞️ Nguồn phụ đề"
            ws["A8"].font = Font(bold=True, size=11)
            ws["A8"].fill = PatternFill("solid", fgColor="FFF3E0")
            ws["B8"] = "auto"
            ws["C8"] = "auto / mota / dan / ocr"
            ws["C8"].font = Font(size=9, italic=True, color="777777")
        return ws

    ws = wb.create_sheet(SHEET_INPUT, 0)
    ws.sheet_properties.tabColor = "C62828"
    for col, w in (("A", 30), ("B", 62), ("C", 42), ("D", 14), ("E", 12)):
        ws.column_dimensions[col].width = w

    ws["A1"] = "🎬 NHẬP LINK VIDEO ➜ TỰ ĐỘNG THÊM TỪ MỚI"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="B71C1C")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 26

    rows = [
        ("🔗 LINK VIDEO  ▸ dán vào đây", "", "Dán link YouTube rồi chạy HA_video.bat"),
        ("🔢 Số từ mới tối đa", 40, "Lấy N từ tần suất cao nhất (0 = lấy hết)"),
        ("✂️ Số ký tự tối thiểu", 2, "2 = bỏ từ đơn 1 chữ; 1 = lấy cả từ 1 chữ"),
        ("🚫 Bỏ hư từ thông dụng", "Có", "Có / Không  (的, 了, 是, 我…)"),
        ("🌐 Ngôn ngữ phụ đề", "auto", "auto  hoặc  zh-Hans / zh-TW"),
        ("📊 TRẠNG THÁI", "— chưa chạy —", "Script tự cập nhật sau mỗi lần chạy"),
        ("🎞️ Nguồn phụ đề", "auto", "auto / mota / dan / ocr"),
    ]
    for i, (lab, val, hint) in enumerate(rows, start=2):
        ws.cell(i, 1, lab).font = Font(bold=True, size=11)
        ws.cell(i, 1).fill = PatternFill("solid", fgColor="FFF3E0")
        ws.cell(i, 2, val)
        ws.cell(i, 3, hint).font = Font(size=9, italic=True, color="777777")
        ws.row_dimensions[i].height = 20
    ws[CELL_LINK].font = Font(size=12, bold=True, color="1565C0")
    ws[CELL_LINK].fill = PatternFill("solid", fgColor="FFFDE7")
    ws.row_dimensions[2].height = 28

    ws["A9"] = "📜 LỊCH SỬ VIDEO ĐÃ XỬ LÝ"
    ws["A9"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A9"].fill = PatternFill("solid", fgColor="1565C0")
    ws.merge_cells("A9:E9")
    for j, h in enumerate(["📅 NGÀY", "🎬 VIDEO", "🔗 LINK", "🆕 TỪ MỚI", "🔁 ĐÃ CÓ"], 1):
        c = ws.cell(10, j, h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="546E7A")
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A11"
    return ws


def ensure_paste_sheet(wb):
    if SHEET_PASTE in wb.sheetnames:
        return wb[SHEET_PASTE]
    ws = wb.create_sheet(SHEET_PASTE, 1)
    ws.sheet_properties.tabColor = "F9A825"
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 40
    ws["A1"] = "📋 DÁN CHỮ HÁN VÀO ĐÂY (dùng khi video không có phụ đề)"
    ws["A1"].font = Font(size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="EF6C00")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24
    ws["A2"] = ("Mỗi dòng 1 câu. Chép từ mô tả video / gõ lại chữ trên màn hình. "
                "Có lẫn tiếng Việt hay pinyin cũng được — script tự lọc.")
    ws["A2"].font = Font(size=9, italic=True, color="777777")
    ws.merge_cells("A2:B2")
    ws["A3"] = "▼ Bắt đầu dán từ dòng 4 ▼"
    ws["A3"].font = Font(bold=True, size=10, color="EF6C00")
    ws.freeze_panes = "A4"
    return ws


def read_config(ws):
    def g(cell, default):
        v = ws[cell].value
        return default if v in (None, "") else v
    link = str(g(CELL_LINK, "")).strip()
    try:
        topn = int(float(g(CELL_TOPN, 40)))
    except Exception:
        topn = 40
    try:
        minch = max(1, int(float(g(CELL_MINCHARS, 2))))
    except Exception:
        minch = 2
    stop = str(g(CELL_STOPWORDS, "Có")).strip().lower() not in (
        "không", "khong", "no", "0", "false", "n")
    lang = str(g(CELL_LANG, "auto")).strip()
    lang = "" if lang.lower() in ("auto", "") else lang
    mode = str(g(CELL_SOURCE, "auto")).strip().lower()
    if mode not in ("auto", "mota", "dan", "ocr"):
        mode = "auto"
    return link, topn, minch, stop, lang, mode


def append_history(ws, title, url, n_new, n_dup):
    r = HISTORY_START_ROW
    while ws.cell(r, 1).value:
        r += 1
    ws.cell(r, 1, datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws.cell(r, 2, title)
    c = ws.cell(r, 3, "▶ Mở video")
    if url:
        c.hyperlink = url
    c.font = Font(color="1565C0", underline="single")
    ws.cell(r, 4, n_new).alignment = Alignment(horizontal="center")
    ws.cell(r, 5, n_dup).alignment = Alignment(horizontal="center")


# ─────────────────────────── Ghi vào sheet từ vựng ────────────────────────────
def last_data_row(ws):
    last = HEADER_ROW
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if ws.cell(r, C_HAN).value:
            last = r
    return last


def clone_style(ws, src_row, dst_row):
    for c in range(1, N_COLS + 1):
        ws.cell(dst_row, c)._style = ws.cell(src_row, c)._style
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height or ROW_HEIGHT


def bump_ranges(ws, old_last, new_last):
    for c in range(1, N_COLS + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.startswith("="):
            def rep(m):
                n = int(m.group(3))
                return (f"{m.group(1)}:{m.group(2)}{max(n, new_last)}"
                        if n >= old_last - 5 else m.group(0))
            ws.cell(1, c).value = re.sub(
                r"(\$?[A-Z]{1,2}\$?\d+):(\$?[A-Z]{1,2}\$?)(\d+)", rep, v)
    try:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(N_COLS)}{new_last}"
    except Exception:
        pass
    try:
        cf = ws.conditional_formatting
        for rng in list(cf._cf_rules.keys()):
            if str(rng.sqref).startswith("H"):
                rules = cf._cf_rules.pop(rng)
                rng.sqref = openpyxl.worksheet.cell_range.MultiCellRange(
                    f"H{FIRST_DATA_ROW}:H{new_last}")
                cf._cf_rules[rng] = rules
    except Exception:
        pass


def write_words(ws, words, video_title, video_url, cache, tpl_row, source=""):
    start = tpl_row + 1
    max_stt = 0
    for r in range(FIRST_DATA_ROW, tpl_row + 1):
        v = ws.cell(r, C_STT).value
        if isinstance(v, (int, float)):
            max_stt = max(max_stt, int(v))
    today = datetime.now().strftime("%a %b %d %Y")

    for i, w in enumerate(words):
        r = start + i
        clone_style(ws, tpl_row, r)
        ws.cell(r, C_STT, max_stt + 1 + i)
        ws.cell(r, C_LEVEL, "VIDEO")
        ws.cell(r, C_HAN, w["han"])
        ws.cell(r, C_PIN, w["pin"])
        ws.cell(r, C_VIET, w["viet"])
        ws.cell(r, C_TRY1, None)
        ws.cell(r, C_TRY2, None)
        ws.cell(r, C_CHECK,
                f'=IF(TRIM($G{r})="","",IF(EXACT(TRIM($G{r}),TRIM($C{r})),"✅ Đúng","❌ Sai"))')
        a = ws.cell(r, C_AUDIO, "▶"); a.hyperlink = audio_url(w["han"])
        y = ws.cell(r, C_YOUG, "🔍"); y.hyperlink = youglish_url(w["han"])
        for c in range(11, 18):
            ws.cell(r, c, None)
        ws.cell(r, C_DATE, today)
        note = f"📹 Từ video · xuất hiện {w['freq']} lần"
        if source:
            note += f" · nguồn: {source}"
        ws.cell(r, C_NOTE, note)
        ws.cell(r, C_CHIET, build_chiet_tu(w["han"], cache))
        t = ws.cell(r, C_TOPIC, f"🎬 {video_title}")
        if video_url:
            t.hyperlink = video_url
        t.font = Font(size=9, bold=True, color="1565C0", underline="single")
        ws.cell(r, C_EX, w.get("ex") or "")
        ws.cell(r, C_EXPIN, w.get("ex_pin") or "")
    return start, start + len(words) - 1


# ────────────────────────────────── MAIN ──────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="YouTube ➜ từ vựng ➜ Excel")
    ap.add_argument("link", nargs="?", help="Link YouTube (bỏ trống = đọc ô B2)")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--text", help="File .txt chứa chữ Hán")
    ap.add_argument("--title", help="Tên video (dùng kèm --text)")
    ap.add_argument("--top", type=int, help="Ghi đè số từ tối đa")
    ap.add_argument("--check", action="store_true", help="Chỉ chẩn đoán video, không ghi")
    ap.add_argument("--ocr", action="store_true", help="Bật OCR chữ nung cứng trên hình")
    ap.add_argument("--mode", choices=["auto", "mota", "dan", "ocr"],
                    help="Ghi đè nguồn phụ đề")
    ap.add_argument("--no-backup", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        sys.exit(f"❌ Không thấy file: {xlsx}")
    if list(xlsx.parent.glob("~$" + xlsx.name)):
        log("⚠️  File có vẻ đang MỞ trong Excel. Hãy ĐÓNG Excel rồi chạy lại.\n")

    log("═" * 62)
    log("  🎬 HA_video v2.0 — YouTube ➜ Từ vựng ➜ Excel")
    log("═" * 62)
    log(f"📂 File: {xlsx.name}")

    wb = openpyxl.load_workbook(xlsx)
    if SHEET_VOCAB not in wb.sheetnames:
        sys.exit(f"❌ Không thấy sheet '{SHEET_VOCAB}'")
    ws = wb[SHEET_VOCAB]
    wsi = ensure_input_sheet(wb)
    ensure_paste_sheet(wb)

    link_cfg, topn, minch, skip_stop, lang, mode = read_config(wsi)
    link = args.link or link_cfg
    if args.top is not None:
        topn = args.top
    if args.mode:
        mode = args.mode
    if args.ocr and mode == "auto":
        pass  # OCR chạy như bước cuối của chuỗi auto

    # ── Chẩn đoán ───────────────────────────────────────────────────────────
    if args.check:
        vid = extract_video_id(link)
        if not vid:
            sys.exit("❌ Cần link video để chẩn đoán.")
        diagnose(vid, wb)
        return

    # ── 1. Lấy chữ Hán ──────────────────────────────────────────────────────
    if args.text:
        raw = Path(args.text).read_text(encoding="utf-8").splitlines()
        lines = clean_lines(raw)
        title = args.title or Path(args.text).stem
        video_url = link or ""
        source = "file text"
        vid = extract_video_id(link) or ""
    else:
        if not link:
            wsi[CELL_STATUS] = "❌ Chưa có link — dán link vào ô B2"
            safe_save(wb, xlsx, quiet=True)
            sys.exit("❌ Chưa có link video.\n   ➜ Mở Excel, sheet '⚙️ NHẬP VIDEO', "
                     "dán link vào ô B2, lưu file rồi chạy lại.")
        vid = extract_video_id(link)
        if not vid:
            sys.exit(f"❌ Link không hợp lệ: {link}")
        video_url = f"https://www.youtube.com/watch?v={vid}"
        log(f"🔗 Video ID: {vid}")
        title = get_video_title(vid)
        log(f"🎬 Tên video: {title}")
        log("⏬ Đang tìm chữ Hán…")
        lines, source, info = gather_chinese(vid, wb, lang, want_ocr=args.ocr, mode=mode)
        if not lines:
            wsi[CELL_STATUS] = "❌ Không tìm được chữ Hán trong video này"
            safe_save(wb, xlsx, quiet=True)
            log("\n❌ Video này không lấy được chữ Hán tự động.")
            try:
                diagnose(vid, wb)
            except Exception:
                pass
            log("\n➜ CÁCH CHẮC ĂN NHẤT:")
            log("   1. Mở file Excel → sheet '📋 DÁN PHỤ ĐỀ'")
            log("   2. Dán/gõ chữ Hán trong video vào cột A (từ dòng 4)")
            log("   3. Lưu, đóng Excel, chạy lại HA_video.bat")
            sys.exit(1)

    log(f"✅ Lấy được {len(lines)} dòng chữ Hán  (nguồn: {source})")

    # ── 2. Tách từ ──────────────────────────────────────────────────────────
    freq, example = segment(lines, minch, skip_stop)
    log(f"🔪 Tách được {len(freq)} từ khác nhau (≥{minch} ký tự"
        f"{', đã bỏ hư từ' if skip_stop else ''})")
    if not freq:
        log("❌ Không có từ nào đạt điều kiện. Thử hạ 'Số ký tự tối thiểu' về 1.")
        return

    # ── 3. Lọc trùng với cột C ──────────────────────────────────────────────
    last = last_data_row(ws)
    existing = {str(ws.cell(r, C_HAN).value).strip()
                for r in range(FIRST_DATA_ROW, last + 1) if ws.cell(r, C_HAN).value}
    log(f"📚 Trong file đang có {len(existing)} từ (đến dòng {last})")

    new_words = [(w, n) for w, n in freq.most_common() if w not in existing]
    n_dup = len(freq) - len(new_words)
    if topn > 0:
        new_words = new_words[:topn]
    log(f"🔁 Bỏ {n_dup} từ đã học · 🆕 Còn {len(new_words)} từ mới sẽ thêm")

    if not new_words:
        wsi[CELL_STATUS] = f"✅ {datetime.now():%H:%M} — video này không có từ mới nào"
        if not args.dry_run:
            safe_save(wb, xlsx)
        log("\n🎉 Không có từ mới — bạn đã biết hết từ trong video này!")
        return

    # ── 4. Dịch nghĩa + câu ví dụ ───────────────────────────────────────────
    log("🌐 Đang dịch nghĩa tiếng Việt…")
    words_only = [w for w, _ in new_words]
    sents = [example.get(w, "") for w, _ in new_words]
    tr_w = gtranslate(words_only)
    tr_s = gtranslate([s for s in sents if s]) if any(sents) else {}

    cache = build_char_cache(ws, last)
    unknown = sorted({ch for w in words_only for ch in w if is_cjk(ch) and ch not in cache})
    if unknown:
        log(f"🧩 Tra nghĩa {len(unknown)} chữ Hán chưa có trong kho chiết tự…")
        tr_c = gtranslate(unknown)
        for ch in unknown:
            cache[ch] = {"hv": "", "mean": tr_c.get(ch, ""), "chiet": ""}
    save_cache(cache)

    rows = []
    for w, n in new_words:
        ex = example.get(w, "")
        ex_pin = ""
        if ex:
            ex_pin = sentence_pinyin(ex)
            v = tr_s.get(ex, "")
            if v:
                ex_pin += f"\n➜ {v}"
        rows.append({"han": w, "pin": to_pinyin(w), "viet": tr_w.get(w, ""),
                     "freq": n, "ex": ex, "ex_pin": ex_pin})

    log("\n" + "-" * 62)
    for i, w in enumerate(rows[:15], 1):
        log(f"  {i:2}. {w['han']:<6} {w['pin']:<16} {w['viet'][:32]:<34} ×{w['freq']}")
    if len(rows) > 15:
        log(f"  … và {len(rows) - 15} từ nữa")
    log("-" * 62 + "\n")

    if args.dry_run:
        log("🔍 --dry-run: KHÔNG ghi file.")
        return

    # ── 5. Ghi vào Excel ────────────────────────────────────────────────────
    if not args.no_backup:
        bak = xlsx.with_name(xlsx.stem + "_backup.xlsx")
        shutil.copy2(xlsx, bak)
        log(f"💾 Đã sao lưu: {bak.name}")

    r1, r2 = write_words(ws, rows, title, video_url, cache, last, source)
    bump_ranges(ws, last, r2)
    append_history(wsi, title, video_url, len(rows), n_dup)
    wsi[CELL_STATUS] = (f"✅ {datetime.now():%Y-%m-%d %H:%M} — thêm {len(rows)} từ mới "
                        f"(dòng {r1}–{r2}), bỏ {n_dup} từ trùng · nguồn: {source}")

    if not safe_save(wb, xlsx):
        sys.exit(1)

    log(f"✅ Đã thêm {len(rows)} từ mới vào dòng {r1}–{r2} của '{SHEET_VOCAB}'")
    log(f"📂 Cột CHỦ ĐỀ = 🎬 {title}")
    log(f"💾 Đã lưu: {xlsx}")
    log("═" * 62)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("\n⛔ Đã huỷ.")
